from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from stock_quant_v2.ops_domain.services.m8_daily_ops_service import M8DailyOpsService
from stock_quant_v2.ops_domain.services.m8_human_review_service import M8HumanReviewService
from stock_quant_v2.ops_domain.services.m8_query_service import M8QueryService
from stock_quant_v2.ops_domain.services.m8_scheduler_service import M8SchedulerService


class M8ExcelExportService:
    def __init__(self, session: Session):
        self.session = session
        self.query_service = M8QueryService(session)
        self.daily_ops_service = M8DailyOpsService(session)
        self.human_review_service = M8HumanReviewService(session)
        self.scheduler_service = M8SchedulerService(session)

    def export_excel_human_review_pack(
        self,
        *,
        output_dir: Path,
        portfolio_id: int,
        profile_code: str | None = None,
    ) -> dict[str, Any]:
        output_dir.mkdir(parents=True, exist_ok=True)

        ops_kpi = self.human_review_service.query_ops_kpi(
            portfolio_id=portfolio_id,
            profile_code=profile_code,
        )

        latest = self.query_service.query_latest_runs(
            portfolio_id=portfolio_id,
            profile_code=profile_code,
        )

        trading_chain = latest.get("trading_chain") or {}
        risk_chain = latest.get("risk_chain") or {}

        paper_chain = None
        snapshot = None
        risk_decision = None
        target_diff = None
        scheduler_health = self.scheduler_service.scheduler_health_check(
            portfolio_id=portfolio_id,
            profile_code=profile_code,
            output_dir=Path("artifacts/m8/daily_ops"),
        )

        if self._has_all(
            trading_chain,
            ["target_run_id", "order_run_id", "fill_run_id", "position_run_id", "snapshot_run_id"],
        ):
            paper_chain = self.query_service.query_paper_chain(
                portfolio_id=portfolio_id,
                target_run_id=int(trading_chain["target_run_id"]),
                order_run_id=int(trading_chain["order_run_id"]),
                fill_run_id=int(trading_chain["fill_run_id"]),
                position_run_id=int(trading_chain["position_run_id"]),
                snapshot_run_id=int(trading_chain["snapshot_run_id"]),
            )
            snapshot = self.query_service.query_portfolio_snapshot(
                portfolio_id=portfolio_id,
                snapshot_run_id=int(trading_chain["snapshot_run_id"]),
            )

        if self._has_all(
            risk_chain,
            ["risk_run_id", "source_target_run_id", "adjusted_target_run_id"],
        ):
            risk_decision = self.query_service.query_risk_decision(
                portfolio_id=portfolio_id,
                source_target_run_id=int(risk_chain["source_target_run_id"]),
                adjusted_target_run_id=int(risk_chain["adjusted_target_run_id"]),
                risk_run_id=int(risk_chain["risk_run_id"]),
                limit=500,
            )
            target_diff = self.query_service.query_target_diff(
                portfolio_id=portfolio_id,
                source_target_run_id=int(risk_chain["source_target_run_id"]),
                adjusted_target_run_id=int(risk_chain["adjusted_target_run_id"]),
                risk_run_id=int(risk_chain["risk_run_id"]),
                limit=500,
            )

        kpi = ops_kpi.get("kpi") or {}
        snapshot_date = kpi.get("snapshot_date") or "latest"
        path = output_dir / f"m8_human_review_pack_p{portfolio_id}_{snapshot_date}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        self._write_title(ws, "M8.9 Human Review Excel Pack", "A1:H1")
        self._write_key_values(
            ws,
            start_row=3,
            title="Overall Status",
            rows=[
                ("portfolio_id", portfolio_id),
                ("profile_code", profile_code),
                ("snapshot_date", kpi.get("snapshot_date")),
                ("ops_kpi_status", ops_kpi.get("overall_status")),
                ("scheduler_status", scheduler_health.get("overall_status")),
                ("scheduler_exit_code", kpi.get("scheduler_exit_code")),
                ("hygiene_status", kpi.get("hygiene_status")),
                ("running_count", kpi.get("running_count")),
            ],
        )

        self._write_key_values(
            ws,
            start_row=13,
            title="Portfolio KPI",
            rows=[
                ("total_equity", kpi.get("total_equity")),
                ("holding_count", kpi.get("holding_count")),
                ("order_count", kpi.get("order_count")),
                ("fill_count", kpi.get("fill_count")),
                ("position_count", kpi.get("position_count")),
                ("snapshot_count", kpi.get("snapshot_count")),
            ],
        )

        self._write_key_values(
            ws,
            start_row=22,
            title="Risk KPI",
            rows=[
                ("risk_decision_count", kpi.get("risk_decision_count")),
                ("risk_pass_count", kpi.get("risk_pass_count")),
                ("risk_warn_count", kpi.get("risk_warn_count")),
                ("risk_reject_count", kpi.get("risk_reject_count")),
                ("risk_adjust_count", kpi.get("risk_adjust_count")),
                ("target_quantity_delta", kpi.get("target_quantity_delta")),
                ("target_amount_delta", kpi.get("target_amount_delta")),
            ],
        )

        self._write_key_values(
            ws,
            start_row=32,
            title="Trading Chain",
            rows=list((latest.get("trading_chain") or {}).items()),
        )

        self._write_key_values(
            ws,
            start_row=41,
            title="Risk Chain",
            rows=list((latest.get("risk_chain") or {}).items()),
        )

        self._write_table(
            wb.create_sheet("Warnings"),
            rows=ops_kpi.get("warnings") or [],
            title="Ops KPI Warnings",
        )

        self._write_table(
            wb.create_sheet("Run Status"),
            rows=ops_kpi.get("run_status_counts") or [],
            title="Run Status Counts",
        )

        self._write_table(
            wb.create_sheet("Risk Reasons"),
            rows=(risk_decision or {}).get("reason_summary") or [],
            title="Risk Reason Summary",
        )

        self._write_table(
            wb.create_sheet("Target Diff"),
            rows=(target_diff or {}).get("diff_rows_preview") or [],
            title="Target Diff Preview",
        )

        self._write_table(
            wb.create_sheet("Paper Chain"),
            rows=[
                {"section": "target", **((paper_chain or {}).get("target") or {})},
                {"section": "order", **((paper_chain or {}).get("order") or {})},
                {"section": "fill", **((paper_chain or {}).get("fill") or {})},
                {"section": "position", **((paper_chain or {}).get("position") or {})},
                {"section": "snapshot", **((paper_chain or {}).get("snapshot") or {})},
            ],
            title="Paper Chain Summary",
        )

        self._write_table(
            wb.create_sheet("Snapshot"),
            rows=[(snapshot or {}).get("snapshot") or {}],
            title="Portfolio Snapshot",
        )

        self._finalize_workbook(wb)
        wb.save(path)

        return {
            "module": "M8.9",
            "query": "export_excel_human_review_pack",
            "portfolio_id": portfolio_id,
            "profile_code": profile_code,
            "file": str(path),
            "ops_kpi_status": ops_kpi.get("overall_status"),
            "scheduler_health_status": scheduler_health.get("overall_status"),
            "overall_status": "PASS"
            if ops_kpi.get("overall_status") in {"PASS", "WARN"}
            and scheduler_health.get("overall_status") == "PASS"
            else "FAIL",
        }

    def export_excel_daily_ops(
        self,
        *,
        output_dir: Path,
        portfolio_id: int,
        profile_code: str | None = None,
    ) -> dict[str, Any]:
        output_dir.mkdir(parents=True, exist_ok=True)

        daily_ops = self.daily_ops_service.daily_ops_check(
            portfolio_id=portfolio_id,
            profile_code=profile_code,
            export_report=True,
            output_dir=Path("artifacts/m8/daily_ops"),
        )

        snapshot_date = (((daily_ops.get("snapshot") or {}).get("snapshot") or {}).get("snapshot_date")) or "latest"
        path = output_dir / f"m8_daily_ops_p{portfolio_id}_{snapshot_date}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Ops"

        self._write_title(ws, "M8.9 Daily Ops Excel Report", "A1:H1")
        self._write_key_values(
            ws,
            start_row=3,
            title="Daily Ops Status",
            rows=[
                ("portfolio_id", portfolio_id),
                ("profile_code", profile_code),
                ("overall_status", daily_ops.get("overall_status")),
                ("checked_at", daily_ops.get("checked_at")),
            ],
        )

        self._write_table(
            wb.create_sheet("Checks"),
            rows=[
                {"check_code": k, **v}
                for k, v in (daily_ops.get("checks") or {}).items()
            ],
            title="Daily Ops Checks",
        )

        self._write_table(
            wb.create_sheet("Warnings"),
            rows=daily_ops.get("warnings") or [],
            title="Daily Ops Warnings",
        )

        self._write_table(
            wb.create_sheet("Failures"),
            rows=daily_ops.get("failures") or [],
            title="Daily Ops Failures",
        )

        self._write_table(
            wb.create_sheet("Risk Decision"),
            rows=[
                {
                    "metric": k,
                    "value": v,
                }
                for k, v in (((daily_ops.get("risk_decision") or {}).get("summary") or {}).items())
            ],
            title="Risk Decision Summary",
        )

        self._write_table(
            wb.create_sheet("Target Diff"),
            rows=[
                {
                    "metric": k,
                    "value": v,
                }
                for k, v in (((daily_ops.get("target_diff") or {}).get("diff_summary") or {}).items())
            ],
            title="Target Diff Summary",
        )

        self._write_table(
            wb.create_sheet("Snapshot"),
            rows=[((daily_ops.get("snapshot") or {}).get("snapshot") or {})],
            title="Portfolio Snapshot",
        )

        self._finalize_workbook(wb)
        wb.save(path)

        return {
            "module": "M8.9",
            "query": "export_excel_daily_ops",
            "portfolio_id": portfolio_id,
            "profile_code": profile_code,
            "file": str(path),
            "daily_ops_status": daily_ops.get("overall_status"),
            "overall_status": "PASS" if daily_ops.get("overall_status") in {"PASS", "WARN"} else "FAIL",
        }

    def export_excel_ops_summary(
        self,
        *,
        output_dir: Path,
        portfolio_id: int,
        profile_code: str | None = None,
    ) -> dict[str, Any]:
        output_dir.mkdir(parents=True, exist_ok=True)

        ops_kpi = self.human_review_service.query_ops_kpi(
            portfolio_id=portfolio_id,
            profile_code=profile_code,
        )

        latest = self.query_service.query_latest_runs(
            portfolio_id=portfolio_id,
            profile_code=profile_code,
        )

        kpi = ops_kpi.get("kpi") or {}
        snapshot_date = kpi.get("snapshot_date") or "latest"
        path = output_dir / f"m8_ops_summary_p{portfolio_id}_{snapshot_date}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Ops Summary"

        self._write_title(ws, "M8.9 Ops Summary Excel Report", "A1:H1")
        self._write_key_values(
            ws,
            start_row=3,
            title="Ops KPI",
            rows=list(kpi.items()),
        )

        self._write_key_values(
            wb.create_sheet("Trading Chain"),
            start_row=1,
            title="Trading Chain",
            rows=list((latest.get("trading_chain") or {}).items()),
        )

        self._write_key_values(
            wb.create_sheet("Risk Chain"),
            start_row=1,
            title="Risk Chain",
            rows=list((latest.get("risk_chain") or {}).items()),
        )

        self._write_table(
            wb.create_sheet("Run Status"),
            rows=ops_kpi.get("run_status_counts") or [],
            title="Run Status Counts",
        )

        self._write_table(
            wb.create_sheet("Warnings"),
            rows=ops_kpi.get("warnings") or [],
            title="Ops KPI Warnings",
        )

        self._finalize_workbook(wb)
        wb.save(path)

        return {
            "module": "M8.9",
            "query": "export_excel_ops_summary",
            "portfolio_id": portfolio_id,
            "profile_code": profile_code,
            "file": str(path),
            "ops_kpi_status": ops_kpi.get("overall_status"),
            "overall_status": "PASS" if ops_kpi.get("overall_status") in {"PASS", "WARN"} else "FAIL",
        }

    def _write_title(self, ws: Any, title: str, merge_range: str) -> None:
        ws.merge_cells(merge_range)
        cell = ws[merge_range.split(":")[0]]
        cell.value = title
        cell.font = Font(name="Calibri", size=16, bold=True, color="1F2937")
        cell.fill = PatternFill("solid", fgColor="EAF2F8")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[cell.row].height = 28

    def _write_key_values(
        self,
        ws: Any,
        *,
        start_row: int,
        title: str,
        rows: list[tuple[str, Any]],
    ) -> None:
        ws.cell(start_row, 1).value = title
        ws.cell(start_row, 1).font = Font(name="Calibri", size=12, bold=True, color="1F2937")
        ws.cell(start_row, 1).fill = PatternFill("solid", fgColor="F3F4F6")

        header_row = start_row + 1
        ws.cell(header_row, 1).value = "metric"
        ws.cell(header_row, 2).value = "value"
        self._style_header_row(ws, header_row, 2)

        for idx, (key, value) in enumerate(rows, start=header_row + 1):
            ws.cell(idx, 1).value = str(key)
            ws.cell(idx, 2).value = self._excel_value(value)
            self._style_body_row(ws, idx, 2)

    def _write_table(self, ws: Any, *, rows: list[dict[str, Any]], title: str) -> None:
        self._write_title(ws, title, "A1:H1")

        if not rows:
            ws.cell(3, 1).value = "No rows"
            return

        fields: list[str] = []
        seen = set()
        for row in rows:
            for key in row.keys():
                if key not in seen:
                    fields.append(key)
                    seen.add(key)

        header_row = 3
        for col_idx, field in enumerate(fields, start=1):
            ws.cell(header_row, col_idx).value = field

        self._style_header_row(ws, header_row, len(fields))

        for row_idx, row in enumerate(rows, start=header_row + 1):
            for col_idx, field in enumerate(fields, start=1):
                ws.cell(row_idx, col_idx).value = self._excel_value(row.get(field))
            self._style_body_row(ws, row_idx, len(fields))

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(fields))}{header_row + len(rows)}"

    def _style_header_row(self, ws: Any, row: int, max_col: int) -> None:
        fill = PatternFill("solid", fgColor="E5E7EB")
        font = Font(name="Calibri", size=10, bold=True, color="111827")
        border = Border(bottom=Side(style="thin", color="D1D5DB"))

        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws.row_dimensions[row].height = 22

    def _style_body_row(self, ws: Any, row: int, max_col: int) -> None:
        border = Border(bottom=Side(style="thin", color="E5E7EB"))
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Calibri", size=10, color="111827")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    def _finalize_workbook(self, wb: Workbook) -> None:
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
            for col in range(1, ws.max_column + 1):
                letter = get_column_letter(col)
                max_len = 10
                for row in range(1, min(ws.max_row, 80) + 1):
                    value = ws.cell(row, col).value
                    if value is not None:
                        max_len = max(max_len, min(len(str(value)), 42))
                ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 44)

            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 18

            ws.freeze_panes = ws.freeze_panes or "A3"

    @staticmethod
    def _excel_value(value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False, default=str)
        return value

    @staticmethod
    def _has_all(payload: dict[str, Any], keys: list[str]) -> bool:
        return all(payload.get(key) is not None for key in keys)